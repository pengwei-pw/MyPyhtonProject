#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @FileName  :Training01.py
# @Time      :2025/07/14 14:01:27
# @Author    :PengWei
import pandas as pd

def demo1():
    path = r'C:\Users\pengw\Desktop\LLM\R06\CD542H\CD542H_R06_LLM_Automation_Test_Report.xlsx'
    df = pd.read_excel(path, sheet_name='result')
    df_1 = df.duplicated(subset=['Query'], keep='first')
    df_1 = df[df_1 == False]
    df_1.to_excel(r"C:\Users\pengw\Desktop\LLM\R06\CD542H\wssssss.xlsx", index=False)

def demo2():
    path_1 = r"C:\Users\pengw\Desktop\LLM\R06\CD542H\wssssss.xlsx"
    df_1 = pd.read_excel(path_1)
    path_2 = r"C:\Users\pengw\Desktop\LLM\llm测试语料\LLM测试用例_全量_持续更新.xlsx"
    df_2 = pd.read_excel(path_2, sheet_name='Sheet1')
    df_2_query = df_2['Query'].tolist()
    for index, row in df_1.iterrows():
        if row['Query'] not in df_2_query:
            print(f"Query '{row['Query']}' not found in df_2")
        
from openpyxl import Workbook
def demo3():
    path1 = r'C:\Users\pengw\Desktop\HALO\LLM\llm测试语料\new0910_PP_003\CX483-H-FAQ.xlsx'
    df1 = pd.read_excel(path1, sheet_name='Sheet1')
    wb = Workbook()
    ws = wb.active
    p_text = 'IGN Run' +'\n'+ 'IVI Power On' +'\n'+ '网络正常'
    e_text = '1.识别正确' +'\n'+ '2.落域正确' +'\n'+ '3.TTS播报正确'
    for index, row in df1.iterrows():
        # ws.merge_cells(start_row=index + 2 + 2*index, start_column=1, end_row=index + 2 + 2*(index+1), end_column=1)
        cell = ws.cell(row=index + 2, column=1)
        cell.value = str(index + 1)
        
        # ws.merge_cells(start_row=index + 2 + 2*index, start_column=2, end_row=index + 2 + 2*(index+1), end_column=2)
        cell = ws.cell(row=index + 2, column=2)
        cell.value = row['Query']

        # ws.merge_cells(start_row=index + 2 + 2*index, start_column=3, end_row=index + 2 + 2*(index+1), end_column=3)
        cell = ws.cell(row=index + 2, column=3)
        if row['二级分类'] in ['用车顾问-FAQ']:
            text = f'按照{row["Query"]}预期结果回答，给出参考答案为：' + '\n' + row["参考答案"]
        elif row['二级分类'] in ['车知识专家-车书', '车知识专家-用车顾问', '车知识专家-网络搜索']:
            text = f'按照{row["Query"]}合理回答即可,给出参考答案为：' + '\n' + row["参考答案"]
        else:
            text = '回答合理即可'
        cell.value = text

        # ws.merge_cells(start_row=index + 2 + 2*index, start_column=4, end_row=index + 2 + 2*(index+1), end_column=4)
        cell = ws.cell(row=index + 2, column=4)
        cell.value = ''

        # ws.merge_cells(start_row=index + 2 + 2*index, start_column=5, end_row=index + 2 + 2*(index+1), end_column=5)
        cell = ws.cell(row=index + 2, column=5)
        cell.value = 'LLM'

        # ws.merge_cells(start_row=index + 2 + 2*index, start_column=6, end_row=index + 2 + 2*(index+1), end_column=6)
        cell = ws.cell(row=index + 2, column=6)
        cell.value = 'CD542MCA_H,CD542MCA_L,CX483'

        # ws.merge_cells(start_row=index + 2 + 2*index, start_column=7, end_row=index + 2 + 2*(index+1), end_column=7)
        cell = ws.cell(row=index + 2, column=7)
        cell.value = ''

        # ws.merge_cells(start_row=index + 2 + 2*index, start_column=8, end_row=index + 2 + 2*(index+1), end_column=8)
        cell = ws.cell(row=index + 2, column=8)
        cell.value = 'Level1'

        # ws.merge_cells(start_row=index + 2 + 2*index, start_column=9, end_row=index + 2 + 2*(index+1), end_column=9)
        cell = ws.cell(row=index + 2, column=9)
        cell.value = 'Medium'

        # ws.merge_cells(start_row=index + 2 + 2*index, start_column=10, end_row=index + 2 + 2*(index+1), end_column=10)
        cell = ws.cell(row=index + 2, column=10)
        cell.value = 'Daily'

        cell = ws.cell(row=index + 2, column=11)
        cell.value = p_text
        
        # ws.merge_cells(start_row=index + 2 + 2*index, start_column=12, end_row=index + 2 + 2*(index+1), end_column=12)
        cell = ws.cell(row=index + 2, column=12)
        cell.value = f'语音唤醒发出指令{row["Query"]}'

        cell = ws.cell(row=index + 2, column=13)
        if row['二级分类'] in ['用车顾问-FAQ']:
            text = '1.识别正确' +'\n'+ '2.落域正确' +'\n'+ '3.TTS播报必须为：{0}'.format(row["参考答案"])
        elif row['二级分类'] in ['车知识专家-车书', '车知识专家-用车顾问', '车知识专家-网络搜索']:
            text = '1.识别正确' +'\n'+ '2.落域正确' +'\n'+ '3.TTS播报合理即可，参考答案：{0}'.format(row["参考答案"])
        else:
            text = '1.识别正确' +'\n'+ '2.落域正确' +'\n'+ '3.TTS播报合理即可'
        cell.value = text

        # ws.merge_cells(start_row=index + 2 + 2*index, start_column=14, end_row=index + 2 + 2*(index+1), end_column=14)
        cell = ws.cell(row=index + 2, column=14)
        if row['二级分类'] in ['车知识专家-车书', '车知识专家-用车顾问', '车知识专家-网络搜索']:
            text = f'车知识Automation'
        elif row['二级分类'] in ['用车顾问-FAQ']:
            text = '车知识FAQAutomation'
        else:
            text = 'AI问答Automation'
        cell.value = text

        # ws.merge_cells(start_row=index + 2 + 2*index, start_column=15, end_row=index + 2 + 2*(index+1), end_column=15)
        cell = ws.cell(row=index + 2, column=15)
        cell.value = ''

        # ws.merge_cells(start_row=index + 2 + 2*index, start_column=16, end_row=index + 2 + 2*(index+1), end_column=16)
        cell = ws.cell(row=index + 2, column=16)
        cell.value = 'Completed'


    wb.save(r'C:\Users\pengw\Desktop\HALO\LLM\llm测试语料\new0910_PP_003\CX483-H-FAQ_testrail_upllm.xlsx')

def ss():
    pd.read_excel(r'C:\Users\pengw\Desktop\HALO\LLM\llm测试语料\new0910_PP_003\CarBit_Performance_Case.xlsx').to_csv(r'C:\Users\pengw\Desktop\HALO\LLM\llm测试语料\new0910_PP_003\CarBit_Performance_Case22222.csv',
            encoding='utf_8_sig',
            index=False,)

if __name__ == "__main__":
    ss()